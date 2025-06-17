import streamlit as st
import pandas as pd
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import io

def data_process(df_raw, mapping_df):
    correction_dict = dict(zip(mapping_df['wrong_address'], mapping_df['right_address']))
    valid_addresses = set(mapping_df['right_address'])

    def correct_address(addr):
        if addr in valid_addresses:
            return addr, "correct", ""
        elif addr in correction_dict:
            return correction_dict[addr], "corrected", ""
        else:
            return addr, "unknown", "未出现在映射表中，请人工确认"

    df_raw[['corrected_address', 'status', 'remark']] = df_raw['address'].apply(
        lambda x: pd.Series(correct_address(x))
    )

    # 写入内存中的 Excel 文件
    output = io.BytesIO()
    df_raw.to_excel(output, index=False)
    output.seek(0)

    # 标红 unknown 行
    wb = load_workbook(output)
    ws = wb.active

    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    # 找到 status 列
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "status":
            status_col = idx
            break

    for row in ws.iter_rows(min_row=2):
        if row[status_col - 1].value == "unknown":
            for cell in row:
                cell.fill = red_fill

    # 保存为最终内存文件
    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output

# === Streamlit 界面 ===
st.title("📍 地址纠错工具")
st.markdown("请上传待判断的地址文件（**需包含一列名为 `address`**）")

uploaded_file = st.file_uploader("上传 Excel 文件（.xlsx）", type=["xlsx"])

if uploaded_file:
    try:
        df_raw = pd.read_excel(uploaded_file)
        mapping_df = pd.read_excel("address_book.xlsx")  # 读取本地映射表

        if 'address' not in df_raw.columns:
            st.error("❌ 上传文件中缺少名为 'address' 的列")
        else:
            if st.button("🚀 开始处理"):
                result_file = data_process(df_raw, mapping_df)
                today_str = date.today().isoformat()
                st.success("✅ 处理完成！点击下方按钮下载结果")
                st.download_button(
                    label="📥 下载处理结果",
                    data=result_file,
                    file_name=f"corrected_{today_str}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    except Exception as e:
        st.error(f"❌ 文件读取失败：{e}")
